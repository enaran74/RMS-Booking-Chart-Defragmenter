#!/usr/bin/env python3
"""
RMS Multi-Property Defragmentation Analyzer - Entry Point
Main orchestration script for analyzing specific properties or all properties
Compatible with Debian 12 Linux Server
"""

import sys
import argparse
import time
import os
from datetime import datetime, date
from typing import List, Dict, Optional
import pandas as pd
from rms_client import RMSClient
from defrag_analyzer import DefragmentationAnalyzer
from excel_generator import ExcelGenerator
from email_sender import EmailSender
from holiday_client import HolidayClient
from utils import setup_logging, get_logger

class MultiPropertyAnalyzer:
    def __init__(self, agent_id: str, agent_password: str, client_id: str, client_password: str, target_property_codes: List[str] = None, enable_emails: bool = False, use_training_db: bool = False):
        """Initialize the multi-property analyzer"""
        
        # Initialize logging
        self.logger = setup_logging()
        self.logger.info("Initializing MultiPropertyAnalyzer")
        self.logger.log_function_entry("MultiPropertyAnalyzer.__init__", 
                                     agent_id=agent_id, 
                                     client_id=client_id, 
                                     target_property_codes=target_property_codes,
                                     enable_emails=enable_emails,
                                     use_training_db=use_training_db)
        
        # Check dependencies
        try:
            import openpyxl
            self.logger.debug("Dependency check: openpyxl available")
        except ImportError:
            error_msg = "Error: openpyxl not available. Excel output is required."
            self.logger.critical(error_msg)
            print("Error: openpyxl not available. Excel output is required.")
            print("Install with: pip install openpyxl")
            sys.exit(1)
        
        # Initialize components with credentials
        self.logger.info("Initializing RMS client")
        self.rms_client = RMSClient(agent_id, agent_password, client_id, client_password, use_training_db=use_training_db)
        self.defrag_analyzer = DefragmentationAnalyzer()
        self.excel_generator = ExcelGenerator()
        self.email_sender = EmailSender()
        
        # Initialize holiday client
        self.logger.info("Initializing holiday client")
        self.holiday_client = HolidayClient()
        
        # Analysis state
        self.all_properties = []
        self.target_properties = []
        self.target_property_codes = target_property_codes or []
        self.enable_emails = enable_emails
        self.use_training_db = use_training_db
        
        # Consolidated analysis data collection
        self.consolidated_daily_data = {}  # {property_code: {date: {'moves': int, 'total_score': float}}}
        self.consolidated_suggestions = []  # List of all suggestions across all properties
        self.consolidated_category_importance_levels = {}  # {property_code: {date: {category: importance_level}}}
        
        # Log configuration
        self.logger.info(f"Configuration: Analysis Period {self.rms_client.constraint_start_date} to {self.rms_client.constraint_end_date}")
        self.logger.info(f"Configuration: Target Properties {self.target_property_codes or 'ALL'}")
        self.logger.info(f"Configuration: Email Notifications {'Enabled' if self.enable_emails else 'Disabled'}")
        self.logger.info(f"Configuration: Database {'Training' if self.use_training_db else 'Live Production'}")
        
        print(f"🚀 RMS MULTI-PROPERTY DEFRAGMENTATION ANALYZER")
        print(f"=" * 60)
        print(f"📅 Analysis Period: {self.rms_client.constraint_start_date} to {self.rms_client.constraint_end_date} (Next 31 days)")
        
        if self.target_property_codes:
            print(f"🎯 Target Properties: {', '.join(self.target_property_codes)}")
        else:
            print(f"🎯 Target Properties: ALL properties")
        
        print(f"🔄 Mode: Real-time API data with comprehensive caching")
        print(f"📧 Email Notifications: {'✅ Enabled' if self.enable_emails else '❌ Disabled'}")
        print(f"🗄️  Database: {'🧪 Training' if self.use_training_db else '🌐 Live Production'}")
        
        self.logger.log_function_exit("MultiPropertyAnalyzer.__init__")

    def run_complete_analysis(self):
        """Main method to run multi-property analysis"""
        start_time = time.time()
        self.logger.log_function_entry("run_complete_analysis")
        
        print(f"\n🎯 STARTING MULTI-PROPERTY ANALYSIS")
        print("=" * 60)
        self.logger.info("Starting multi-property analysis")
        
        # Clear any existing cache for fresh start
        self.logger.info("Clearing cache for fresh start")
        self.rms_client.clear_cache()
        
        # Step 1: Authenticate and discover properties
        self.logger.info("Step 1: Authenticating with RMS API")
        if not self.rms_client.authenticate():
            self.logger.error("Authentication failed - cannot proceed")
            print("❌ Authentication failed - cannot proceed")
            return False
        
        # Get discovered properties and filter valid ones
        self.logger.info("Fetching all properties from RMS")
        self.all_properties = self.rms_client.get_all_properties()
        self.logger.log_data_summary("Properties discovered", len(self.all_properties))
        
        # Filter properties based on target codes
        if self.target_property_codes:
            self.logger.info(f"Filtering properties by target codes: {self.target_property_codes}")
            self.target_properties = self._filter_properties_by_codes(self.all_properties)
        else:
            self.logger.info("Filtering all properties for valid ones")
            self.target_properties = self._filter_valid_properties(self.all_properties)
        
        if not self.target_properties:
            self.logger.error("No valid properties found - cannot proceed")
            print("❌ No valid properties found - cannot proceed")
            return False
        
        self.logger.log_data_summary("Properties after filtering", len(self.target_properties), f"from {len(self.all_properties)} total")
        print(f"📊 Filtered to {len(self.target_properties)} valid properties (from {len(self.all_properties)} total)")
        
        # Step 2: Run analysis for each property
        self.logger.info("Step 2: Running multi-property analysis")
        result = self.run_multi_property_analysis()
        
        duration = time.time() - start_time
        self.logger.log_performance_metric("Complete analysis", duration)
        self.logger.log_function_exit("run_complete_analysis", result)
        
        return result
    
    def _filter_valid_properties(self, properties: List[Dict]) -> List[Dict]:
        """Filter properties to only include those with valid IDs, names, and active status"""
        valid_properties = []
        
        print(f"\n🔍 DEBUGGING PROPERTY VALIDATION:")
        print(f"Total properties to validate: {len(properties)}")
        
        for i, prop in enumerate(properties[:5]):  # Show first 5 for debugging
            print(f"Property {i+1}: {prop}")
        
        if len(properties) > 5:
            print(f"... and {len(properties) - 5} more properties")
        
        for prop in properties:
            property_id = prop.get('id')
            property_name = prop.get('name')
            property_inactive = prop.get('inactive', False)
            
            print(f"   Checking: ID={property_id}, Name={property_name}, Inactive={property_inactive}")
            
            # Skip properties with invalid or missing data
            if (property_id is None or 
                property_name is None or 
                str(property_id).strip() == '' or 
                str(property_name).strip() == ''):
                print(f"   ❌ Skipped: Invalid data")
                continue
            
            # Skip inactive properties
            if property_inactive:
                print(f"   ❌ Skipped: Inactive property")
                continue
            
            # Try to convert property_id to int to ensure it's valid
            try:
                int(property_id)
                valid_properties.append(prop)
                print(f"   ✅ Valid: {property_name} (ID: {property_id})")
            except (ValueError, TypeError):
                print(f"   ❌ Skipped: Invalid ID format")
                continue
        
        print(f"\n📊 Validation Summary:")
        print(f"   Total: {len(properties)}")
        print(f"   Valid: {len(valid_properties)}")
        print(f"   Invalid: {len(properties) - len(valid_properties)}")
        
        return valid_properties
    
    def _filter_properties_by_codes(self, properties: List[Dict]) -> List[Dict]:
        """Filter properties to only include those matching target codes"""
        filtered_properties = []
        
        print(f"\n🔍 FILTERING PROPERTIES BY CODES:")
        print(f"Target codes: {', '.join(self.target_property_codes)}")
        
        # Show first few properties for debugging
        print(f"\n🔍 DEBUGGING PROPERTY CODES:")
        for i, prop in enumerate(properties[:5]):
            code = prop.get('code', 'NO_CODE')
            name = prop.get('name', 'NO_NAME')
            print(f"   Property {i+1}: Code='{code}', Name='{name}'")
        
        for prop in properties:
            property_id = prop.get('id')
            property_name = prop.get('name')
            property_code = prop.get('code', '').upper()
            property_inactive = prop.get('inactive', False)
            
            # Skip properties with invalid or missing data
            if (property_id is None or 
                property_name is None or 
                str(property_id).strip() == '' or 
                str(property_name).strip() == ''):
                continue
            
            # Skip inactive properties
            if property_inactive:
                continue
            
            # Clean property code (remove hyphens and normalize)
            clean_property_code = property_code.replace('-', '').replace('_', '').strip()
            
            # Check if this property matches any target code
            target_codes_clean = [code.upper().replace('-', '').replace('_', '').strip() for code in self.target_property_codes]
            
            if clean_property_code in target_codes_clean:
                try:
                    int(property_id)  # Ensure ID is valid
                    filtered_properties.append(prop)
                    print(f"   ✅ Matched: {property_name} (Code: {property_code} -> {clean_property_code}, ID: {property_id})")
                except (ValueError, TypeError):
                    print(f"   ❌ Skipped: Invalid ID format for {property_name}")
                    continue
            else:
                # Debug output for first few non-matches
                if len(filtered_properties) < 3:
                    print(f"   ❌ No match: {property_name} (Code: {property_code} -> {clean_property_code})")
        
        print(f"\n📊 Code Filtering Summary:")
        print(f"   Total properties: {len(properties)}")
        print(f"   Matched properties: {len(filtered_properties)}")
        print(f"   Unmatched properties: {len(properties) - len(filtered_properties)}")
        
        if len(filtered_properties) == 0:
            print(f"\n⚠️  NO MATCHES FOUND!")
            print(f"   Target codes (cleaned): {target_codes_clean}")
            print(f"   Available codes (first 10): {[prop.get('code', 'NO_CODE') for prop in properties[:10]]}")
        
        return filtered_properties
    
    def run_multi_property_analysis(self):
        """Process all properties with progress tracking and user control"""
        total_properties = len(self.target_properties)
        successful_analyses = 0
        failed_analyses = 0
        
        print(f"\n🏢 MULTI-PROPERTY PROCESSING")
        print("=" * 50)
        print(f"📊 Total Properties to Analyze: {total_properties}")
        print(f"📅 Analysis Period: {self.rms_client.constraint_start_date} to {self.rms_client.constraint_end_date}")
        
        # Show initial overall progress
        self._print_overall_parks_progress(0, total_properties)
        
        for property_index, property_info in enumerate(self.target_properties, 1):
            property_id = property_info.get('id')
            property_name = property_info.get('name', f'Property-{property_id}')
            
            # Ensure property_id is an integer
            try:
                property_id = int(property_id)
            except (ValueError, TypeError):
                print(f"⚠️  Skipping property with invalid ID: {property_id}")
                failed_analyses += 1
                continue
            
            try:
                # Analyze this property
                success = self._analyze_single_property(property_id, property_name, property_index, total_properties)
                
                if success:
                    successful_analyses += 1
                    print(f"✅ Property analysis completed successfully")
                else:
                    failed_analyses += 1
                    print(f"❌ Property analysis failed")
                    
            except Exception as e:
                failed_analyses += 1
                print(f"💥 Property analysis error: {e}")
            
            # Update and show overall progress
            self._print_overall_parks_progress(property_index, total_properties)
            
            # Show completion status
            print(f"\n📊 Progress Summary:")
            print(f"   ✅ Completed: {successful_analyses}")
            print(f"   ❌ Failed: {failed_analyses}")
            print(f"   📈 Remaining: {total_properties - property_index}")
        
        # Final progress update
        self._print_overall_parks_progress(total_properties, total_properties)
        
        # Cache performance summary
        cache_stats = self.rms_client.get_cache_stats()
        print(f"\n💾 CACHE PERFORMANCE SUMMARY:")
        print("=" * 40)
        print(f"📊 Categories cached: {cache_stats['categories_cache_size']}")
        print(f"🏠 Areas cached: {cache_stats['areas_cache_size']}")
        print(f"🏢 Properties cached: {cache_stats['total_cached_properties']}")
        print(f"⏱️  Cache TTL: {cache_stats['cache_ttl_seconds']} seconds")
        
        # Email performance summary (only if emails were enabled)
        if self.enable_emails:
            self.email_sender.print_email_summary()
        else:
            print(f"\n📧 EMAIL SUMMARY: Not applicable (emails disabled)")
        
        # Endpoint usage summary
        self.rms_client.print_endpoint_summary()
        
        # Generate consolidated Excel file
        if successful_analyses > 0:
            self._generate_consolidated_excel()
        
        # Final summary
        print(f"\n🎉 MULTI-PROPERTY ANALYSIS COMPLETE!")
        print("=" * 60)
        print(f"📊 Total Properties: {total_properties}")
        print(f"✅ Successful: {successful_analyses}")
        print(f"❌ Failed: {failed_analyses}")
        print(f"📈 Success Rate: {(successful_analyses/total_properties)*100:.1f}%")
        
        return successful_analyses > 0
    
    def _get_property_email(self, property_id: int) -> str:
        """Get email address for a specific property"""
        all_properties = self.rms_client.get_all_properties()
        for prop in all_properties:
            if prop.get('id') == property_id:
                return prop.get('email')
        return None
    

    
    def _analyze_single_property(self, property_id: int, property_name: str, property_index: int, total_properties: int) -> bool:
        """Analyze a single property"""
        start_time = time.time()
        self.logger.log_function_entry("_analyze_single_property", 
                                     property_id=property_id, 
                                     property_name=property_name, 
                                     property_index=property_index, 
                                     total_properties=total_properties)
        
        # Find the property info to get the code
        property_info = None
        for prop in self.target_properties:
            if prop.get('id') == property_id:
                property_info = prop
                break
        
        # Get property code for filename, fallback to ID if no code
        property_code = property_info.get('code', str(property_id)) if property_info else str(property_id)
        
        # Clean property code for filename (remove trailing hyphens, underscores, spaces)
        clean_property_code = property_code.rstrip('- _')
        
        # Generate output filename using cleaned property code (will overwrite existing files)
        # Use OUTPUT_DIR environment variable or fallback to current directory
        output_dir = os.environ.get('OUTPUT_DIR', os.path.join(os.getcwd(), 'output'))
        os.makedirs(output_dir, exist_ok=True)
        output_filename = os.path.join(output_dir, f"{clean_property_code}-Defragmentation-Analysis.xlsx")
        
        # Log the output directory being used
        self.logger.info(f"Using output directory: {os.path.abspath(output_dir)}")
        print(f"📁 Output directory: {os.path.abspath(output_dir)}")
        
        # Log property analysis start
        self.logger.log_property_analysis_start(clean_property_code, property_id, property_name)
        
        print(f"\n🏢 ANALYZING PROPERTY: {property_name}")
        print("=" * 60)
        print(f"🆔 Property ID: {property_id}")
        print(f"🏷️  Property Code: {property_code}")
        print(f"📁 Output File: {output_filename}")
        
        try:
            # Step 1: Fetch live inventory data
            self.logger.info(f"Step 1: Fetching inventory data for {property_name}")
            inventory_df = self.rms_client.fetch_inventory_data(property_id, property_name)
            if inventory_df.empty:
                self.logger.warning(f"No inventory data retrieved for {property_name} - skipping property")
                print("❌ No inventory data retrieved - skipping property")
                return False
            
            self.logger.log_data_summary("Inventory units", len(inventory_df), f"for {property_name}")
            
            # Step 2: Fetch live reservations data
            self.logger.info(f"Step 2: Fetching reservations data for {property_name}")
            reservations_df = self.rms_client.fetch_reservations_data(property_id, property_name)
            if reservations_df.empty:
                self.logger.warning(f"No reservations data retrieved for {property_name} - skipping property")
                print("❌ No reservations data retrieved - skipping property")
                return False
            
            self.logger.log_data_summary("Reservations", len(reservations_df), f"for {property_name}")
            
            # Step 3: Perform defragmentation analysis
            self.logger.info(f"Step 3: Performing defragmentation analysis for {property_name}")
            analysis_start = time.time()
            suggestions = self.defrag_analyzer.analyze_defragmentation(
                reservations_df, 
                inventory_df, 
                self.rms_client.constraint_start_date, 
                self.rms_client.constraint_end_date
            )
            analysis_duration = time.time() - analysis_start
            self.logger.log_performance_metric("Defragmentation analysis", analysis_duration, f"for {property_name}")
            self.logger.log_data_summary("Move suggestions", len(suggestions), f"for {property_name}")
            
            # Step 3.5: Collect daily move data for consolidated analysis
            self.logger.debug(f"Collecting daily move data for {property_name}")
            self._collect_daily_move_data(clean_property_code, suggestions, reservations_df, inventory_df)
            
            # Step 3.6: Collect suggestions for consolidated analysis
            self.logger.debug(f"Collecting suggestions data for {property_name}")
            self._collect_suggestions_data(property_id, property_name, property_code, suggestions)
            
            # Step 3.7: Perform holiday analysis
            self.logger.info(f"Step 3.7: Performing holiday analysis for {property_name}")
            holiday_start = time.time()
            
            # Get state code for holiday analysis
            state_code = self.rms_client.get_property_state_code(property_id)
            if not state_code:
                self.logger.warning(f"Could not determine state code for {property_name} - skipping holiday analysis")
                print("⚠️  Could not determine state code - skipping holiday analysis")
                holiday_suggestions = []
                holiday_data = {'holiday_periods': []}
            else:
                # Debug: Force refresh cache for this state to ensure fresh data
                self.holiday_client.force_refresh_cache_for_state(state_code)
                print(f"🗺️  Detected state code: {state_code} for {property_name}")
                self.holiday_client.debug_cache_contents(state_code)
                # Fetch combined holiday periods (public + school holidays) for 2-month forward analysis
                holiday_periods = self.holiday_client.get_combined_holiday_periods_2month_forward(
                    state_code,
                    date.today()  # Start from today
                )
                
                if holiday_periods:
                    # Perform 2-month forward holiday defragmentation analysis
                    holiday_suggestions = self.defrag_analyzer.analyze_holiday_defragmentation_2month_forward(
                        reservations_df,
                        inventory_df,
                        holiday_periods,
                        self.rms_client.constraint_start_date,
                        self.rms_client.constraint_end_date
                    )
                    
                    # Merge regular and holiday suggestions
                    merged_suggestions = self.defrag_analyzer.merge_move_lists(suggestions, holiday_suggestions)
                    
                    holiday_data = {
                        'holiday_periods': holiday_periods,
                        'state_code': state_code
                    }
                    
                    # Count school vs public holidays
                    school_holidays = [h for h in holiday_periods if h.get('is_school_holiday')]
                    public_holidays = [h for h in holiday_periods if not h.get('is_school_holiday')]
                    
                    print(f"🎄 2-Month Forward Combined Holiday Analysis: {len(holiday_periods)} periods ({len(school_holidays)} school, {len(public_holidays)} public), {len(holiday_suggestions)} holiday moves")
                    print(f"📋 Total Merged Suggestions: {len(merged_suggestions)} moves")
                    
                    # Update suggestions with merged list
                    suggestions = merged_suggestions
                else:
                    self.logger.info(f"No holiday periods found for {property_name} in {state_code} in 2-month forward window")
                    print(f"📅 No holiday periods found for {state_code} in 2-month forward window")
                    holiday_suggestions = []
                    holiday_data = {'holiday_periods': []}
            
            holiday_duration = time.time() - holiday_start
            self.logger.log_performance_metric("Holiday analysis", holiday_duration, f"for {property_name}")
            self.logger.log_data_summary("Holiday suggestions", len(holiday_suggestions), f"for {property_name}")
            
            # Step 4: Create holiday-enhanced Excel output
            self.logger.info(f"Step 4: Creating holiday-enhanced Excel output for {property_name}")
            excel_start = time.time()
            
            # Separate regular and holiday suggestions for Excel generation
            regular_suggestions = [s for s in suggestions if not s.get('is_holiday_move', False)]
            holiday_suggestions = [s for s in suggestions if s.get('is_holiday_move', False)]
            
            excel_success, category_importance_levels = self.excel_generator.create_holiday_enhanced_excel(
                reservations_df, 
                inventory_df,
                regular_suggestions,
                holiday_suggestions,
                property_id,
                property_name,
                holiday_data,
                self.rms_client.constraint_start_date,
                self.rms_client.constraint_end_date,
                output_filename
            )
            
            excel_duration = time.time() - excel_start
            self.logger.log_performance_metric("Excel generation", excel_duration, f"for {property_name}")
            
            # Step 4.5: Collect category importance levels for consolidated analysis
            self.logger.debug(f"Collecting category importance levels for {property_name}")
            self._collect_category_importance_levels(clean_property_code, category_importance_levels)
            
            # Step 5: Send email with Excel attachment (if enabled)
            email_success = False
            if self.enable_emails:
                self.logger.info(f"Step 5: Sending email for {property_name}")
                # Get property's email address
                property_email = self._get_property_email(property_id)
                
                email_success = self.email_sender.send_holiday_enhanced_email(
                    property_name,
                    property_id,
                    output_filename,
                    regular_suggestions,
                    holiday_suggestions,
                    holiday_data,
                    self.rms_client.constraint_start_date,
                    self.rms_client.constraint_end_date,
                    excel_success,
                    property_email,
                    self.use_training_db
                )
                self.logger.log_email_operation("Property analysis", property_email or "unknown", email_success)
            else:
                self.logger.debug(f"Email notifications disabled - skipping email for {property_name}")
                print(f"📧 Email notifications disabled - skipping email for {property_name}")
            
            # Step 6: Display results summary
            self._display_property_summary(suggestions, excel_success, property_id, property_name, output_filename, holiday_suggestions, holiday_data)
            
            # Show cache and email performance
            cache_stats = self.rms_client.get_cache_stats()
            self.logger.log_data_summary("Cached properties", cache_stats['total_cached_properties'])
            print(f"💾 Cache Performance: {cache_stats['total_cached_properties']} properties cached")
            if self.enable_emails:
                print(f"📧 Email Status: {'✅ Sent' if email_success else '❌ Failed'}")
            else:
                print(f"📧 Email Status: Disabled")
            
            # Log property analysis completion
            total_duration = time.time() - start_time
            self.logger.log_performance_metric("Property analysis", total_duration, f"for {property_name}")
            self.logger.log_property_analysis_complete(clean_property_code, len(suggestions), output_filename)
            self.logger.log_function_exit("_analyze_single_property", excel_success)
            
            return excel_success
            
        except Exception as e:
            self.logger.log_error_with_context(e, f"Property analysis for {property_name}")
            print(f"💥 Error analyzing property {property_name}: {e}")
            return False
    
    def _print_overall_parks_progress(self, completed: int, total: int):
        """Print overall parks progress bar in red"""
        bar_length = 60
        progress = completed / total if total > 0 else 0
        filled_length = int(bar_length * progress)
        bar = '█' * filled_length + '-' * (bar_length - filled_length)
        percent = progress * 100
        
        # Red progress bar using ANSI color codes
        red_start = '\033[91m'  # Bright red
        red_end = '\033[0m'     # Reset color
        
        print(f"\n🔴 OVERALL PARKS PROGRESS:")
        print(f"{red_start}[{bar}] {percent:5.1f}% - Completed: {completed}/{total} Parks{red_end}")
    
    def _collect_daily_move_data(self, property_code: str, suggestions: List[Dict], reservations_df: pd.DataFrame, inventory_df: pd.DataFrame):
        """Collect daily move opportunities data for consolidated analysis"""
        from datetime import timedelta
        
        # Initialize property data
        self.consolidated_daily_data[property_code] = {}
        
        # Generate date range
        date_range = []
        current = self.rms_client.constraint_start_date
        while current <= self.rms_client.constraint_end_date:
            date_range.append(current)
            current += timedelta(days=1)
        
        # Initialize daily data
        for date in date_range:
            self.consolidated_daily_data[property_code][date] = {'moves': 0, 'total_score': 0.0}
        
        # Add move opportunities to daily data
        for suggestion in suggestions:
            res_no = suggestion['Reservation_No']
            res_row = reservations_df[reservations_df['Res No'] == res_no]
            if not res_row.empty:
                res = res_row.iloc[0]
                arrive_date = self._parse_date(res['Arrive'])
                depart_date = self._parse_date(res['Depart'])
                
                if arrive_date and depart_date:
                    # Convert datetime to date for comparison
                    arrive_date = arrive_date.date()
                    depart_date = depart_date.date()
                    current_date = max(arrive_date, self.rms_client.constraint_start_date)
                    end_date = min(depart_date, self.rms_client.constraint_end_date + timedelta(days=1))
                    
                    while current_date < end_date and current_date <= self.rms_client.constraint_end_date:
                        if current_date in self.consolidated_daily_data[property_code]:
                            self.consolidated_daily_data[property_code][current_date]['moves'] += 1
                            self.consolidated_daily_data[property_code][current_date]['total_score'] += suggestion['Improvement_Score']
                        current_date += timedelta(days=1)
        

    
    def _collect_suggestions_data(self, property_id: int, property_name: str, property_code: str, suggestions: List[Dict]):
        """Collect suggestions data for consolidated analysis"""
        # Clean property code by removing trailing hyphens
        clean_property_code = property_code.rstrip('-')
        
        for suggestion in suggestions:
            # Create a copy of the suggestion with property information
            consolidated_suggestion = suggestion.copy()
            consolidated_suggestion['PropertyId'] = property_id
            consolidated_suggestion['PropertyCode'] = clean_property_code
            consolidated_suggestion['PropertyName'] = property_name
            self.consolidated_suggestions.append(consolidated_suggestion)
    
    def _collect_category_importance_levels(self, property_code: str, category_importance_levels: dict):
        """Collect category importance levels from individual property analysis"""
        # Clean property code by removing trailing hyphens
        clean_property_code = property_code.rstrip('-')
        
        # Store the category importance levels for this property
        self.consolidated_category_importance_levels[clean_property_code] = category_importance_levels
    
    def _parse_date(self, date_str: str):
        """Parse date string to datetime object"""
        try:
            # Try DD/MM/YYYY HH:MM format first (e.g., "24/08/2025 14:00")
            return datetime.strptime(date_str, '%d/%m/%Y %H:%M')
        except (ValueError, TypeError):
            try:
                # Try YYYY-MM-DD format as fallback
                return datetime.strptime(date_str, '%Y-%m-%d')
            except (ValueError, TypeError):
                return None
    
    def _importance_level_rank(self, level: str) -> int:
        """Get rank of importance level for comparison (higher number = higher importance)"""
        ranks = {"None": 0, "Low": 1, "Medium": 2, "High": 3}
        return ranks.get(level, 0)
    

    
    def _generate_consolidated_excel(self):
        """Generate consolidated Excel file with all properties' daily move data"""
        from datetime import timedelta
        
        if not self.consolidated_daily_data:
            print("⚠️  No consolidated data available - skipping consolidated Excel generation")
            return
        
        print(f"\n📊 GENERATING CONSOLIDATED EXCEL FILE")
        print("=" * 50)
        
        # Generate date range
        date_range = []
        current = self.rms_client.constraint_start_date
        while current <= self.rms_client.constraint_end_date:
            date_range.append(current)
            current += timedelta(days=1)
        
        # Sort property codes alphabetically
        property_codes = sorted(self.consolidated_daily_data.keys())
        
        # Create workbook
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Consolidated Daily Moves"
        
        # Colors for heatmap
        def get_heatmap_color(score: float, max_score: float) -> PatternFill:
            if max_score == 0:
                hex_color = "FFE6E6"  # Very light red
            else:
                intensity = min(score / max_score, 1.0)
                if intensity < 0.01:
                    hex_color = "FFE6E6"  # Very light red
                elif intensity < 0.3:
                    hex_color = "FFB3B3"  # Light red
                elif intensity < 0.6:
                    hex_color = "FF6666"  # Medium red
                else:
                    hex_color = "CC0000"  # Dark red
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        
        # Calculate max score for color scaling
        max_score = 0
        for property_code in property_codes:
            for date in date_range:
                if date in self.consolidated_daily_data[property_code]:
                    max_score = max(max_score, self.consolidated_daily_data[property_code][date]['total_score'])
        
        if max_score == 0:
            max_score = 1  # Prevent division by zero
        
        # Headers
        current_row = 1
        
        # Title
        ws.cell(row=current_row, column=1, value="CONSOLIDATED DAILY MOVE OPPORTUNITIES").font = Font(bold=True, size=14, color="000080")
        ws.merge_cells(f'A{current_row}:{get_column_letter(len(date_range) + 1)}{current_row}')
        current_row += 2
        
        # Date headers
        ws.cell(row=current_row, column=1, value="Property Code").font = Font(bold=True)
        for col_idx, date in enumerate(date_range, start=2):
            date_cell = ws.cell(row=current_row, column=col_idx, value=date.strftime("%d/%m"))
            date_cell.font = Font(bold=True)
            date_cell.alignment = Alignment(horizontal="center")
        current_row += 1
        
        # Property rows with move data - two rows per property
        for property_code in property_codes:
            # Clean property code for display (remove trailing hyphens)
            clean_property_code = property_code.rstrip('-')
            
            # Row 1: Number of moves
            ws.cell(row=current_row, column=1, value=f"{clean_property_code} Moves").font = Font(bold=True)
            
            for col_idx, date in enumerate(date_range, start=2):
                if date in self.consolidated_daily_data[property_code]:
                    data = self.consolidated_daily_data[property_code][date]
                    cell = ws.cell(row=current_row, column=col_idx, value=data['moves'])
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True)
                    
                    # Add tooltip for moves row
                    if data['moves'] > 0:
                        from openpyxl.comments import Comment
                        comment_text = (f"Property: {clean_property_code}\n"
                                      f"Date: {date.strftime('%d/%m/%Y')}\n"
                                      f"Moves available: {data['moves']}\n"
                                      f"Total opportunity: {data['total_score']:.2f}\n"
                                      f"Avg move value: {data['total_score']/data['moves']:.2f}")
                        comment = Comment(comment_text, "DefragScript")
                        comment.width = 180
                        comment.height = 100
                        cell.comment = comment
                else:
                    ws.cell(row=current_row, column=col_idx, value=0)
            
            current_row += 1
            
            # Row 2: Move importance level (with heatmap coloring)
            ws.cell(row=current_row, column=1, value=f"{clean_property_code} Move Importance").font = Font(bold=True)
            
            for col_idx, date in enumerate(date_range, start=2):
                if date in self.consolidated_daily_data[property_code]:
                    data = self.consolidated_daily_data[property_code][date]
                    
                    # Get the highest importance level for this property/date from pre-calculated data
                    importance_level = "None"
                    if property_code in self.consolidated_category_importance_levels:
                        if date in self.consolidated_category_importance_levels[property_code]:
                            category_levels = self.consolidated_category_importance_levels[property_code][date].values()
                            if category_levels:
                                # Apply hierarchical logic: High > Medium > Low > None
                                if "High" in category_levels:
                                    importance_level = "High"
                                elif "Medium" in category_levels:
                                    importance_level = "Medium"
                                elif "Low" in category_levels:
                                    importance_level = "Low"
                    
                    # Set cell value and color based on importance level
                    cell = ws.cell(row=current_row, column=col_idx, value=importance_level)
                    
                    # Color coding based on importance level
                    if importance_level == "High":
                        cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")  # Dark red
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif importance_level == "Medium":
                        cell.fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")  # Medium red
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif importance_level == "Low":
                        cell.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")  # Light red
                        cell.font = Font(color="000000", bold=True)
                    else:  # None
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Very light red
                        cell.font = Font(color="000000", bold=True)
                    
                    cell.alignment = Alignment(horizontal="center")
                    
                    # Add tooltip for importance row
                    if importance_level != "None":
                        from openpyxl.comments import Comment
                        comment_text = (f"Property: {clean_property_code}\n"
                                      f"Date: {date.strftime('%d/%m/%Y')}\n"
                                      f"Move Importance: {importance_level}\n"
                                      f"Moves available: {data['moves']}\n"
                                      f"Categories with moves: {', '.join(self.consolidated_category_importance_levels[property_code][date].keys())}")
                        comment = Comment(comment_text, "DefragScript")
                        comment.width = 180
                        comment.height = 100
                        cell.comment = comment
                else:
                    ws.cell(row=current_row, column=col_idx, value="None")
                    ws.cell(row=current_row, column=col_idx).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            current_row += 1
        
        # Legend
        current_row += 1
        ws.cell(row=current_row, column=1, value="MOVE IMPORTANCE LEGEND").font = Font(bold=True, size=12)
        current_row += 1
        
        legend_items = [
            ("None", "FFE6E6", "No move opportunities or sufficient availability"),
            ("Low", "FFB3B3", "Low importance moves available"),
            ("Medium", "FF6666", "Medium importance moves available"),
            ("High", "CC0000", "High importance moves available")
        ]
        
        for label, color, description in legend_items:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value="").fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=current_row, column=3, value=description)
            current_row += 1
        
        # Add scoring explanation
        current_row += 1
        ws.cell(row=current_row, column=1, value="MOVE IMPORTANCE LOGIC EXPLANATION").font = Font(bold=True, size=12)
        current_row += 1
        
        scoring_explanation = [
            "MOVE IMPORTANCE LEVELS:",
            "• Based on strategic importance of categories with move opportunities on each date",
            "• Each category's importance is calculated using fragmentation, availability density, and unit scale",
            "• Importance levels: None (0), Low (<0.33), Medium (0.33-0.67), High (>0.67)",
            "",
            "HIERARCHICAL LOGIC:",
            "• High: If any category has High importance for that night",
            "• Medium: If no High, but at least one Medium importance for that night",
            "• Low: If no High/Medium, but at least one Low importance for that night",
            "• None: If all categories have None importance for that night",
            "",
            "COLOR CODING:",
            "• Dark Red = High importance moves available",
            "• Medium Red = Medium importance moves available",
            "• Light Red = Low importance moves available",
            "• Very Light Red = No move opportunities or sufficient availability",
            "• Tooltips show categories with moves when hovering over cells"
        ]
        
        for explanation in scoring_explanation:
            cell = ws.cell(row=current_row, column=1, value=explanation)
            if explanation.endswith(":"):
                cell.font = Font(bold=True)
            else:
                cell.font = Font(color="000000")
            ws.merge_cells(f'A{current_row}:{get_column_letter(len(date_range) + 1)}{current_row}')
            current_row += 1
        
        # Add Suggested_Moves sheet if we have consolidated suggestions
        if self.consolidated_suggestions:
            self._create_suggested_moves_sheet(wb)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file in output directory
        output_dir = os.environ.get('OUTPUT_DIR', os.path.join(os.getcwd(), 'output'))
        os.makedirs(output_dir, exist_ok=True)
        output_filename = os.path.join(output_dir, "Full_Defragmentation_Analysis.xlsx")
        
        # Log the output directory being used
        self.logger.info(f"Using output directory for consolidated file: {os.path.abspath(output_dir)}")
        print(f"📁 Consolidated output directory: {os.path.abspath(output_dir)}")
        
        wb.save(output_filename)
        
        print(f"✅ Consolidated Excel file created: {output_filename}")
        print(f"📊 Properties included: {len(property_codes)}")
        print(f"📅 Date range: {date_range[0].strftime('%d/%m/%Y')} to {date_range[-1].strftime('%d/%m/%Y')}")
        print(f"🎨 Move Importance: Color-coded levels (High/Medium/Low/None) showing strategic move opportunities")
        if self.consolidated_suggestions:
            print(f"📋 Suggested_Moves sheet: {len(self.consolidated_suggestions)} total moves across all properties")
    
    def _create_suggested_moves_sheet(self, wb):
        """Create Suggested_Moves sheet with all suggestions across all properties"""
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        ws = wb.create_sheet("Suggested_Moves")
        
        # Define headers
        headers = [
            "PropertyId", "PropertyCode", "PropertyName", "Move Order", "Reservation No", 
            "Guest Surname", "Current Unit", "Suggested Unit", "Category", "Status", 
            "Arrive Date", "Depart Date", "Nights", "Improvement Score", "Reason"
        ]
        
        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        normal_font = Font(color="000000")
        
        # Create headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Sort suggestions by property code, then by move order
        sorted_suggestions = sorted(
            self.consolidated_suggestions, 
            key=lambda x: (x['PropertyCode'], x['Sequential_Order'])
        )
        
        # Add data rows
        for row_idx, suggestion in enumerate(sorted_suggestions, start=2):
            data = [
                suggestion['PropertyId'],
                suggestion['PropertyCode'],
                suggestion['PropertyName'],
                suggestion['Sequential_Order'],
                suggestion['Reservation_No'],
                suggestion['Surname'],
                suggestion['Current_Unit'],
                suggestion['Suggested_Unit'],
                suggestion['Category'],
                suggestion['Status'],
                suggestion['Arrive_Date'],
                suggestion['Depart_Date'],
                suggestion['Nights'],
                suggestion['Improvement_Score'],
                suggestion['Reason']
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                
                # Center align numeric and date columns
                if col_idx in [1, 4, 5, 11, 12, 13, 14]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # Set column widths
        column_widths = [12, 15, 25, 12, 15, 15, 18, 18, 35, 12, 12, 12, 8, 15, 60]
        for col_idx, width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
        
        # Add summary section
        summary_start_row = len(sorted_suggestions) + 4
        summary_info = [
            "CONSOLIDATED SUGGESTIONS SUMMARY:",
            f"Total Moves Suggested: {len(sorted_suggestions)}",
            f"Properties Analyzed: {len(set(s['PropertyCode'] for s in sorted_suggestions))}",
            f"Analysis Period: {self.rms_client.constraint_start_date} to {self.rms_client.constraint_end_date}",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "IMPLEMENTATION NOTES:",
            "• Apply moves in property and category-based order (PropertyCode, then Move Order)",
            "• Each move considers the effect of previous moves within the same property",
            "• Only reservations entirely within analysis period are included",
            "• Fixed reservations are excluded from suggestions",
            "• Higher improvement scores = better defragmentation results",
            "• Real-time data from RMS API at time of analysis"
        ]
        
        for i, info in enumerate(summary_info):
            cell = ws.cell(row=summary_start_row + i, column=1, value=info)
            if info.endswith(":"):
                cell.font = Font(bold=True)
            else:
                cell.font = normal_font
            ws.merge_cells(f'A{summary_start_row + i}:O{summary_start_row + i}')
    
    def _display_property_summary(self, suggestions, excel_success: bool, property_id: int, property_name: str, output_filename: str, holiday_suggestions: List[Dict] = None, holiday_data: Dict = None):
        """Display summary for single property with holiday information"""
        print(f"\n📋 PROPERTY ANALYSIS SUMMARY")
        print("=" * 40)
        print(f"🏢 Property: {property_name} (ID: {property_id})")
        print(f"📅 Period: {self.rms_client.constraint_start_date} to {self.rms_client.constraint_end_date}")
        print(f"🔄 Data Source: {'🧪 Training' if self.use_training_db else '🌐 Live'} RMS API")
        
        # Separate regular and holiday suggestions
        regular_suggestions = [s for s in suggestions if not s.get('is_holiday_move', False)]
        holiday_suggestions = holiday_suggestions or [s for s in suggestions if s.get('is_holiday_move', False)]
        
        print(f"📊 Regular Suggestions: {len(regular_suggestions)} optimization moves")
        print(f"🎄 Holiday Suggestions: {len(holiday_suggestions)} holiday-specific moves")
        print(f"📋 Total Suggestions: {len(suggestions)} moves")
        
        # Show holiday periods if available
        if holiday_data and 'holiday_periods' in holiday_data and holiday_data['holiday_periods']:
            print(f"📅 Holiday Periods: {len(holiday_data['holiday_periods'])} periods analyzed")
            for period in holiday_data['holiday_periods'][:3]:  # Show first 3
                print(f"   • {period['name']} ({period['importance']} importance)")
            if len(holiday_data['holiday_periods']) > 3:
                print(f"   • ... and {len(holiday_data['holiday_periods']) - 3} more")
        
        if excel_success:
            print(f"✅ Excel Report: {output_filename}")
        else:
            print(f"❌ Excel output failed")
        
        if suggestions:
            print(f"\n🔝 TOP 3 MOVE RECOMMENDATIONS:")
            print("-" * 80)
            print(f"{'Order':<8} {'Type':<6} {'Guest':<12} {'From':<15} {'To':<15}")
            print("-" * 80)
            
            for suggestion in suggestions[:3]:
                move_type = "🎄" if suggestion.get('is_holiday_move', False) else "📋"
                order = suggestion.get('move_id', suggestion.get('Sequential_Order', ''))
                guest = suggestion.get('guest_name', suggestion.get('Surname', ''))[:11]
                from_unit = suggestion.get('from_unit', suggestion.get('Current_Unit', ''))[:14]
                to_unit = suggestion.get('to_unit', suggestion.get('Suggested_Unit', ''))[:14]
                
                print(f"{order:<8} {move_type:<6} {guest:<12} {from_unit:<15} {to_unit:<15}")
        else:
            print(f"\n✅ No optimization needed - reservations are well organized!")

def parse_arguments():
    """Parse command line arguments with environment variable support"""
    parser = argparse.ArgumentParser(
        description='RMS Multi-Property Defragmentation Analyzer (Debian 12 Linux Compatible)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 start.py --agent-id ***REMOVED*** --agent-password "********" --client-id ***REMOVED*** --client-password "********" -p SADE,QROC,TCRA    # Analyze specific properties
  python3 start.py --agent-id ***REMOVED*** --agent-password "********" --client-id ***REMOVED*** --client-password "********" -p ALL               # Analyze all properties
  python3 start.py --agent-id ***REMOVED*** --agent-password "********" --client-id ***REMOVED*** --client-password "********" -p SADE              # Analyze single property
  python3 start.py --agent-id ***REMOVED*** --agent-password "********" --client-id ***REMOVED*** --client-password "********" -e -p ALL            # Analyze all properties with email notifications
  python3 start.py --agent-id ***REMOVED*** --agent-password "********" --client-id ***REMOVED*** --client-password "********" -e -p SADE,QROC      # Analyze specific properties with email notifications
  python3 start.py --agent-id ***REMOVED*** --agent-password "********" --client-id ***REMOVED*** --client-password "********" -t -p ALL            # Analyze all properties using training database
  python3 start.py --agent-id ***REMOVED*** --agent-password "********" --client-id ***REMOVED*** --client-password "********" -t -e -p SADE,QROC   # Analyze specific properties with emails using training database (emails sent to operations@discoveryparks.com.au)

Environment Variables (for Linux server deployment):
  AGENT_ID, AGENT_PASSWORD, CLIENT_ID, CLIENT_PASSWORD - RMS API credentials
  TARGET_PROPERTIES - Property codes to analyze (comma-separated) or ALL
  ENABLE_EMAILS - Set to 'true' to enable email notifications
  USE_TRAINING_DB - Set to 'true' to use training database

Note: All RMS credentials are MANDATORY. Replace the example credentials with your actual RMS credentials.
        """
    )
    
    parser.add_argument(
        '--agent-id',
        type=str,
        required=False,
        default=os.environ.get('AGENT_ID'),
        help='RMS Agent ID (MANDATORY, can also be set via AGENT_ID environment variable)'
    )
    
    parser.add_argument(
        '--agent-password',
        type=str,
        required=False,
        default=os.environ.get('AGENT_PASSWORD'),
        help='RMS Agent Password (MANDATORY, can also be set via AGENT_PASSWORD environment variable)'
    )
    
    parser.add_argument(
        '--client-id',
        type=str,
        required=False,
        default=os.environ.get('CLIENT_ID'),
        help='RMS Client ID (MANDATORY, can also be set via CLIENT_ID environment variable)'
    )
    
    parser.add_argument(
        '--client-password',
        type=str,
        required=False,
        default=os.environ.get('CLIENT_PASSWORD'),
        help='RMS Client Password (MANDATORY, can also be set via CLIENT_PASSWORD environment variable)'
    )
    
    parser.add_argument(
        '-p', '--properties',
        type=str,
        required=False,
        default=os.environ.get('TARGET_PROPERTIES', 'ALL'),
        help='Property codes to analyze (comma-separated) or ALL for all properties (can also be set via TARGET_PROPERTIES environment variable)'
    )
    
    parser.add_argument(
        '-e', '--email',
        action='store_true',
        default=os.environ.get('ENABLE_EMAILS', 'false').lower() == 'true',
        help='Send email notifications with Excel attachments after each property analysis (can also be set via ENABLE_EMAILS=true environment variable)'
    )
    
    parser.add_argument(
        '-t', '--training',
        action='store_true',
        default=os.environ.get('USE_TRAINING_DB', 'false').lower() == 'true',
        help='Use training database instead of live production data (can also be set via USE_TRAINING_DB=true environment variable)'
    )
    

    

    

    
    # Custom error handler to show examples
    def error_handler(message):
        print("🚀 RMS MULTI-PROPERTY DEFRAGMENTATION ANALYZER (Debian 12 Linux)")
        print("=" * 70)
        print("❌ ERROR: Missing required arguments")
        print()
        print("📋 USAGE EXAMPLES:")
        print("  python3 start.py --agent-id ***REMOVED*** --agent-password \"********\" --client-id ***REMOVED*** --client-password \"********\" -p SADE,QROC,TCRA    # Analyze specific properties")
        print("  python3 start.py --agent-id ***REMOVED*** --agent-password \"********\" --client-id ***REMOVED*** --client-password \"********\" -p ALL               # Analyze all properties")
        print("  python3 start.py --agent-id ***REMOVED*** --agent-password \"********\" --client-id ***REMOVED*** --client-password \"********\" -p SADE              # Analyze single property")
        print("  python3 start.py --agent-id ***REMOVED*** --agent-password \"********\" --client-id ***REMOVED*** --client-password \"********\" -e -p ALL            # Analyze all properties with email notifications")
        print("  python3 start.py --agent-id ***REMOVED*** --agent-password \"********\" --client-id ***REMOVED*** --client-password \"********\" -t -p ALL            # Analyze all properties using training database")
        print()
        print("🌐 LINUX SERVER DEPLOYMENT (using environment variables):")
        print("  export AGENT_ID=***REMOVED***")
        print("  export AGENT_PASSWORD=\"********\"")
        print("  export CLIENT_ID=***REMOVED***")
        print("  export CLIENT_PASSWORD=\"********\"")
        print("  export TARGET_PROPERTIES=ALL")
        print("  export ENABLE_EMAILS=true")
        print("  export USE_TRAINING_DB=false")
        print("  python3 start.py")
        print()
        print("📁 CONFIGURATION FILE (for Linux server):")
        print("  Edit /etc/bookingchart-defragmenter/config.env")
        print("  Then run: /opt/bookingchart-defragmenter/run_defragmentation.sh")
        print()
        print("Note: All RMS credentials are MANDATORY. Replace the example credentials with your actual RMS credentials.")
        print()
        print("📖 For more help:")
        print("  python3 start.py -h")
        print()
        sys.exit(1)
    
    parser.error = error_handler
    
    return parser.parse_args()

def main():
    """Main execution function for multi-property analysis"""
    
    # Initialize logging at the start
    logger = setup_logging()
    logger.info("Application started")
    logger.log_function_entry("main")
    
    # Parse command line arguments
    args = parse_arguments()
    
    # Validate required credentials
    if not args.agent_id:
        logger.error("Agent ID is required (--agent-id or AGENT_ID environment variable)")
        print("❌ Error: Agent ID is required (--agent-id or AGENT_ID environment variable)")
        sys.exit(1)
    
    if not args.agent_password:
        logger.error("Agent password is required (--agent-password or AGENT_PASSWORD environment variable)")
        print("❌ Error: Agent password is required (--agent-password or AGENT_PASSWORD environment variable)")
        sys.exit(1)
    
    if not args.client_id:
        logger.error("Client ID is required (--client-id or CLIENT_ID environment variable)")
        print("❌ Error: Client ID is required (--client-id or CLIENT_ID environment variable)")
        sys.exit(1)
    
    if not args.client_password:
        logger.error("Client password is required (--client-password or CLIENT_PASSWORD environment variable)")
        print("❌ Error: Client password is required (--client-password or CLIENT_PASSWORD environment variable)")
        sys.exit(1)
    
    # Check if properties argument is required for analysis
    if not args.properties:
        logger.error("Properties argument (-p) is required for analysis")
        print("❌ Error: Properties argument (-p) is required for analysis")
        sys.exit(1)
    
    # Process property codes
    if args.properties.upper() == 'ALL':
        target_codes = []
        logger.info("Mode: Analyzing ALL properties")
        print(f"🎯 MODE: Analyzing ALL properties")
    else:
        target_codes = [code.strip().upper() for code in args.properties.split(',')]
        logger.info(f"Mode: Analyzing specific properties: {', '.join(target_codes)}")
        print(f"🎯 MODE: Analyzing specific properties: {', '.join(target_codes)}")
    
    # Initialize analyzer with credentials, target codes, email setting, and training database flag
    logger.info("Initializing MultiPropertyAnalyzer")
    analyzer = MultiPropertyAnalyzer(args.agent_id, args.agent_password, args.client_id, args.client_password, target_codes, args.email, args.training)
    success = analyzer.run_complete_analysis()
    
    if success:
        logger.info("Multi-property defragmentation analysis completed successfully")
        print(f"\n🎊 SUCCESS: Multi-property defragmentation analysis complete!")
    else:
        logger.error("Multi-property analysis encountered errors")
        print(f"\n💥 FAILED: Multi-property analysis encountered errors")
    
    logger.log_function_exit("main", success)
    return success

if __name__ == "__main__":
    main()