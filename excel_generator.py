#!/usr/bin/env python3
"""
Excel Generator
Creates Excel workbooks with visual charts and suggestions tables
Compatible with Debian 12 Linux Server
"""

import pandas as pd
import time
import os
from datetime import datetime, timedelta
from collections import defaultdict
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from utils import get_logger

class ExcelGenerator:
    def __init__(self):
        """Initialize the Excel generator"""
        self.logger = get_logger()
        self.logger.debug("ExcelGenerator initialized")

    def create_excel_output(self, reservations_df: pd.DataFrame, inventory_df: pd.DataFrame, 
                          suggestions: List[Dict], property_id: int, property_name: str,
                          constraint_start_date, constraint_end_date, output_filename: str):
        """Create Excel workbook with visual chart and suggestions for specified property"""
        start_time = time.time()
        self.logger.log_function_entry("create_excel_output", 
                                     property_id=property_id, 
                                     property_name=property_name,
                                     output_filename=output_filename,
                                     suggestions_count=len(suggestions))
        
        print(f"\n📊 CREATING EXCEL OUTPUT")
        print("=" * 40)
        print(f"📁 File: {output_filename}")
        
        self.logger.info(f"Creating Excel output for {property_name}: {output_filename}")
        
        wb = Workbook()
        
        # Sheet 1: Visual Chart
        chart_sheet = wb.active
        chart_sheet.title = "Visual Chart"
        
        # Sheet 2: Move Suggestions
        suggestions_sheet = wb.create_sheet("Move Suggestions")
        
        # Create both sheets and get category importance levels
        self.logger.info("Creating Visual Chart sheet")
        category_importance_levels = self._create_visual_chart_sheet(
            chart_sheet, reservations_df, inventory_df, suggestions, 
            constraint_start_date, constraint_end_date
        )
        
        self.logger.info("Creating Move Suggestions sheet")
        self._create_suggestions_table_sheet(
            suggestions_sheet, suggestions, property_id, property_name,
            constraint_start_date, constraint_end_date
        )
        
        try:
            # Ensure output directory exists (for Linux server deployment)
            output_dir = os.path.dirname(output_filename)
            if output_dir and not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir, mode=0o755, exist_ok=True)
                    self.logger.info(f"Created output directory: {output_dir}")
                except PermissionError:
                    self.logger.warning(f"Could not create output directory: {output_dir}, using current directory")
                    output_filename = os.path.basename(output_filename)
            
            wb.save(output_filename)
            duration = time.time() - start_time
            self.logger.log_performance_metric("Excel file creation", duration, f"for {property_name}")
            self.logger.log_excel_generation(output_filename, ["Visual Chart", "Move Suggestions"])
            print(f"✅ Excel workbook saved successfully!")
            print(f"   📋 Sheet 1: Visual Chart - Daily heatmap + booking grid")
            print(f"   📋 Sheet 2: Move Suggestions - Detailed recommendations table")
            print(f"   📁 File location: {os.path.abspath(output_filename)}")
            self.logger.log_function_exit("create_excel_output", True)
            return True, category_importance_levels
        except Exception as e:
            self.logger.log_error_with_context(e, f"Excel file creation for {property_name}")
            print(f"❌ Error saving Excel file: {e}")
            return False, {}
    
    def _create_visual_chart_sheet(self, ws, reservations_df: pd.DataFrame, inventory_df: pd.DataFrame, 
                                 suggestions: List[Dict], constraint_start_date, constraint_end_date):
        """Create visual booking chart with daily opportunity heatmap"""
        
        date_range = []
        current_date = constraint_start_date
        while current_date <= constraint_end_date:
            date_range.append(current_date)
            current_date += timedelta(days=1)
        
        # Build suggestions lookup with move direction
        suggestions_lookup = {}
        for suggestion in suggestions:
            # Get current unit from the reservation
            res_no = suggestion['Reservation_No']
            res_row = reservations_df[reservations_df['Res No'] == res_no]
            current_unit = res_row.iloc[0]['Unit/Site'] if not res_row.empty else None
            
            suggestions_lookup[suggestion['Reservation_No']] = {
                'order': suggestion['Sequential_Order'],
                'target_unit': suggestion['Suggested_Unit'],
                'current_unit': current_unit,
                'score': suggestion['Improvement_Score']
            }
        
        # Calculate daily opportunities for heatmap
        daily_opportunities = {}
        for date in date_range:
            daily_opportunities[date] = {'moves': 0, 'total_score': 0.0}
        
        # Calculate category-specific daily opportunities with strategic importance weighting
        category_daily_opportunities = {}
        for date in date_range:
            category_daily_opportunities[date] = {}
        
        # Calculate strategic importance for each category
        from defrag_analyzer import DefragmentationAnalyzer
        analyzer = DefragmentationAnalyzer()
        category_importance = analyzer.calculate_category_strategic_importance(
            reservations_df, inventory_df, constraint_start_date, constraint_end_date
        )
        
        # Add move opportunities to daily data with strategic weighting
        for suggestion in suggestions:
            res_no = suggestion['Reservation_No']
            res_row = reservations_df[reservations_df['Res No'] == res_no]
            if not res_row.empty:
                res = res_row.iloc[0]
                arrive_date = self._parse_date(res['Arrive'])
                depart_date = self._parse_date(res['Depart'])
                category = res['Category']
                
                # Get strategic importance weight for this category
                strategic_weight = category_importance.get(category, 1.0)
                
                if arrive_date and depart_date:
                    current_date = max(arrive_date, constraint_start_date)
                    end_date = min(depart_date, constraint_end_date + timedelta(days=1))
                    
                    while current_date < end_date and current_date <= constraint_end_date:
                        if current_date in daily_opportunities:
                            daily_opportunities[current_date]['moves'] += 1
                            daily_opportunities[current_date]['total_score'] += suggestion['Improvement_Score']
                            
                            # Add to category-specific data with strategic weighting
                            if category not in category_daily_opportunities[current_date]:
                                category_daily_opportunities[current_date][category] = {'moves': 0, 'total_score': 0.0, 'strategic_score': 0.0}
                            
                            category_daily_opportunities[current_date][category]['moves'] += 1
                            category_daily_opportunities[current_date][category]['total_score'] += suggestion['Improvement_Score']
                            # Apply strategic weighting to the score
                            category_daily_opportunities[current_date][category]['strategic_score'] += suggestion['Improvement_Score'] * strategic_weight
                        current_date += timedelta(days=1)
        
        # Colors and fonts
        colors = {
            'header': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'category': PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
            'arrived': PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),
            'confirmed': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
            'unconfirmed': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            'maintenance': PatternFill(start_color="800080", end_color="800080", fill_type="solid"),
            'pencil': PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid"),
            'move_suggestion': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        }
        
        # Border styles
        borders = {
            'reservation': Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
        }
        
        fonts = {
            'header': Font(color="FFFFFF", bold=True),
            'category': Font(bold=True),
            'normal': Font(color="000000"),
            'move': Font(color="FFFFFF", bold=True)
        }
        
        # DAILY OPPORTUNITY HEATMAP
        current_row = 1
        
        ws.cell(row=current_row, column=1, value="DAILY MOVE OPPORTUNITIES").font = Font(bold=True, size=14, color="000080")
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2
        
        # Heatmap headers
        ws.cell(row=current_row, column=1, value="Date").font = Font(bold=True)
        ws.cell(row=current_row + 1, column=1, value="Move Count").font = Font(bold=True)
        
        # Calculate max score for move count color scaling
        max_score = max([data['total_score'] for data in daily_opportunities.values()]) if daily_opportunities else 1
        
        # Create heatmap data
        for col_idx, date in enumerate(date_range, start=2):
            data = daily_opportunities[date]
            
            # Date header
            date_cell = ws.cell(row=current_row, column=col_idx, value=date.strftime("%d/%m"))
            date_cell.font = Font(bold=True)
            date_cell.alignment = Alignment(horizontal="center")
            
            # Move count with color
            moves_cell = ws.cell(row=current_row + 1, column=col_idx, value=data['moves'])
            moves_cell.fill = self._get_heatmap_color(data['total_score'], max_score)
            moves_cell.alignment = Alignment(horizontal="center")
            moves_cell.font = Font(color="FFFFFF" if data['total_score'] > max_score * 0.5 else "000000", bold=True)
            
            # Add tooltip for moves
            if data['moves'] > 0:
                comment_text = (f"Date: {date.strftime('%d/%m/%Y')}\n"
                              f"Moves available: {data['moves']}\n"
                              f"Total opportunity: {data['total_score']:.2f}\n"
                              f"Avg move value: {data['total_score']/data['moves']:.2f}")
                comment = Comment(comment_text, "DefragScript")
                comment.width = 180
                comment.height = 80
                moves_cell.comment = comment
        
        current_row += 2
        
        # Add category-specific move scores
        # Get all unique categories that have moves
        all_categories = set()
        for date_data in category_daily_opportunities.values():
            all_categories.update(date_data.keys())
        
        if all_categories:
            # Sort categories for consistent display
            sorted_categories = sorted(all_categories)
            
            for category in sorted_categories:
                    # Category label
                    ws.cell(row=current_row, column=1, value=f"{category} Importance").font = Font(bold=True, color="000080")
                    
                    # Category strategic scores for each date
                    for col_idx, date in enumerate(date_range, start=2):
                        category_data = category_daily_opportunities[date].get(category, {'moves': 0, 'total_score': 0.0, 'strategic_score': 0.0})
                        
                        # Use strategic score instead of raw total score
                        strategic_score = category_data.get('strategic_score', 0.0)
                        
                        # Convert strategic score to importance level and display
                        importance_level = self._get_importance_level(strategic_score)
                        score_cell = ws.cell(row=current_row, column=col_idx, value=importance_level)
                        score_cell.alignment = Alignment(horizontal="center")
                        
                        # Apply color coding based on importance level
                        if strategic_score > 0:
                            if importance_level == "High":
                                score_cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
                                score_cell.font = Font(color="FFFFFF", bold=True)
                            elif importance_level == "Medium":
                                score_cell.fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
                                score_cell.font = Font(color="000000", bold=True)
                            elif importance_level == "Low":
                                score_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                                score_cell.font = Font(color="000000", bold=True)
                            
                            # Convert strategic score to Low/Medium/High category
                            importance_level = self._get_importance_level(strategic_score)
                            
                            # Add tooltip for category strategic scores
                            comment_text = (f"Category: {category}\n"
                                          f"Date: {date.strftime('%d/%m/%Y')}\n"
                                          f"Moves available: {category_data['moves']}\n"
                                          f"Importance: {importance_level}\n"
                                          f"Strategic score: {strategic_score:.2f}\n"
                                          f"Raw opportunity: {category_data['total_score']:.2f}\n"
                                          f"Strategic importance: {category_importance.get(category, 1.0):.2f}")
                            comment = Comment(comment_text, "DefragScript")
                            comment.width = 220
                            comment.height = 120
                            score_cell.comment = comment
                        else:
                            score_cell.font = Font(color="000000")
                    
                    current_row += 1
        
        current_row += 2
        
        # BOOKING CHART
        ws.cell(row=current_row, column=1, value="BOOKING CHART").font = Font(bold=True, size=14, color="000080")
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 2
        
        booking_chart_start_row = current_row
        
        # Headers
        ws.cell(row=booking_chart_start_row, column=1, value="Unit/Site").fill = colors['header']
        ws.cell(row=booking_chart_start_row, column=1).font = fonts['header']
        
        for col_idx, date in enumerate(date_range, start=2):
            cell = ws.cell(row=booking_chart_start_row, column=col_idx, value=date.strftime("%d/%m"))
            cell.fill = colors['header']
            cell.font = fonts['header']
            cell.alignment = Alignment(horizontal="center")
        
        # Build occupancy matrix
        occupancy = {}
        for _, res in reservations_df.iterrows():
            arrive_date = self._parse_date(res['Arrive'])
            depart_date = self._parse_date(res['Depart'])
            unit = res['Unit/Site']
            is_fixed = self._is_reservation_fixed(res)
            
            if arrive_date and depart_date:
                current_date = max(arrive_date, constraint_start_date)
                end_date = min(depart_date, constraint_end_date + timedelta(days=1))
                
                while current_date < end_date and current_date <= constraint_end_date:
                    occupancy[(unit, current_date)] = {
                        'res_no': res['Res No'],
                        'surname': res['Surname'],
                        'status': res['Status'],
                        'category': res['Category'],
                        'arrive': arrive_date,
                        'depart': depart_date,
                        'nights': res['Nights'],
                        'fixed': is_fixed
                    }
                    current_date += timedelta(days=1)
        
        # Group units by category
        units_by_category = defaultdict(list)
        for _, row in inventory_df.iterrows():
            units_by_category[row['Category']].append(row['Unit/Site'])
        
        # Create unit position mapping for move direction calculation
        unit_positions = {}
        current_row = booking_chart_start_row + 1
        
        # Calculate unit positions in the chart
        for category in sorted(units_by_category.keys()):
            units = sorted(units_by_category[category])
            current_row += 1  # Skip category header
            
            for unit in units:
                unit_positions[unit] = current_row
                current_row += 1
        
        current_row = booking_chart_start_row + 1
        
        # Create booking grid
        for category in sorted(units_by_category.keys()):
            units = sorted(units_by_category[category])
            
            # Category header - merge across entire chart
            category_cell = ws.cell(row=current_row, column=1, value=category)
            category_cell.fill = colors['category']
            category_cell.font = fonts['category']
            category_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Merge category header across all date columns
            end_col = len(date_range) + 1
            if end_col > 1:
                start_col_letter = get_column_letter(1)
                end_col_letter = get_column_letter(end_col)
                merge_range = f"{start_col_letter}{current_row}:{end_col_letter}{current_row}"
                ws.merge_cells(merge_range)
                # Re-apply formatting to the merged cell
                category_cell = ws.cell(row=current_row, column=1)
                category_cell.fill = colors['category']
                category_cell.font = fonts['category']
                category_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            current_row += 1
            
            # Units in category
            for unit in units:
                unit_cell = ws.cell(row=current_row, column=1, value=unit)
                unit_cell.font = fonts['normal']
                unit_cell.alignment = Alignment(wrap_text=True)
                
                # Group consecutive dates for the same reservation to create merged cells
                col_idx = 2
                while col_idx <= len(date_range) + 1:
                    current_date = date_range[col_idx - 2]
                    
                    if (unit, current_date) in occupancy:
                        res_info = occupancy[(unit, current_date)]
                        res_no = res_info['res_no']
                        surname = res_info['surname']
                        status = res_info['status']
                        is_fixed = res_info['fixed']
                        
                        # Find the span of this reservation (consecutive dates)
                        span_start = col_idx
                        span_end = col_idx
                        
                        # Check if this is a suggested move
                        if res_no in suggestions_lookup:
                            # For move suggestions, find the full span like regular reservations
                            suggestion = suggestions_lookup[res_no]
                            span_start = col_idx
                            span_end = col_idx
                            
                            # Find the span of this move suggestion (consecutive dates)
                            while span_end <= len(date_range) + 1:
                                next_date = date_range[span_end - 2] if span_end <= len(date_range) else None
                                if next_date and (unit, next_date) in occupancy:
                                    next_res_info = occupancy[(unit, next_date)]
                                    if (next_res_info['res_no'] == res_no and 
                                        next_res_info['status'] == status and 
                                        next_res_info['fixed'] == is_fixed):
                                        span_end += 1
                                    else:
                                        break
                                else:
                                    break
                            
                            # Create merged cell for the move suggestion span
                            if span_end > span_start:
                                # Apply borders to all cells in the range BEFORE merging
                                for col in range(span_start, span_end):
                                    border_cell = ws.cell(row=current_row, column=col)
                                    border_cell.border = borders['reservation']
                                
                                # Merge cells horizontally
                                start_col_letter = get_column_letter(span_start)
                                end_col_letter = get_column_letter(span_end - 1)
                                merge_range = f"{start_col_letter}{current_row}:{end_col_letter}{current_row}"
                                ws.merge_cells(merge_range)
                                
                                # Apply formatting to the merged cell
                                cell = ws.cell(row=current_row, column=span_start)
                                cell.fill = colors['move_suggestion']
                                cell.font = fonts['normal']  # Use same font as regular bookings for consistent alignment
                                
                                # Add directional arrow before move number
                                direction_icon = self._get_move_direction_icon(
                                    suggestion['current_unit'], 
                                    suggestion['target_unit'], 
                                    unit_positions
                                )
                                cell.value = f"{direction_icon} {suggestion['order']}"
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                                # Enhanced comment with span information
                                span_nights = span_end - span_start
                                comment_text = (f"MOVE #{suggestion['order']}\n"
                                              f"Res: {res_no}\n"
                                              f"Guest: {surname}\n"
                                              f"Move to: {suggestion['target_unit']}\n"
                                              f"Score: {suggestion['score']}\n"
                                              f"Span: {span_nights} nights")
                                comment = Comment(comment_text, "DefragScript")
                                comment.width = 200
                                comment.height = 120
                                cell.comment = comment
                                
                                col_idx = span_end
                            else:
                                # Single night move suggestion (no merge needed)
                                cell = ws.cell(row=current_row, column=col_idx)
                                cell.fill = colors['move_suggestion']
                                cell.font = fonts['normal']  # Use same font as regular bookings for consistent alignment
                                
                                # Add directional arrow before move number
                                direction_icon = self._get_move_direction_icon(
                                    suggestion['current_unit'], 
                                    suggestion['target_unit'], 
                                    unit_positions
                                )
                                cell.value = f"{direction_icon} {suggestion['order']}"
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                cell.border = borders['reservation']
                                
                                comment_text = (f"MOVE #{suggestion['order']}\n"
                                              f"Res: {res_no}\n"
                                              f"Guest: {surname}\n"
                                              f"Move to: {suggestion['target_unit']}\n"
                                              f"Score: {suggestion['score']}")
                                comment = Comment(comment_text, "DefragScript")
                                comment.width = 200
                                comment.height = 100
                                cell.comment = comment
                                
                                col_idx += 1
                            
                        else:
                            # For regular reservations, find the full span
                            while span_end <= len(date_range) + 1:
                                next_date = date_range[span_end - 2] if span_end <= len(date_range) else None
                                if next_date and (unit, next_date) in occupancy:
                                    next_res_info = occupancy[(unit, next_date)]
                                    if (next_res_info['res_no'] == res_no and 
                                        next_res_info['status'] == status and 
                                        next_res_info['fixed'] == is_fixed):
                                        span_end += 1
                                    else:
                                        break
                                else:
                                    break
                            
                            # Create merged cell for the reservation span
                            if span_end > span_start:
                                # Apply borders to all cells in the range BEFORE merging
                                for col in range(span_start, span_end):
                                    border_cell = ws.cell(row=current_row, column=col)
                                    border_cell.border = borders['reservation']
                                
                                # Merge cells horizontally
                                start_col_letter = get_column_letter(span_start)
                                end_col_letter = get_column_letter(span_end - 1)
                                merge_range = f"{start_col_letter}{current_row}:{end_col_letter}{current_row}"
                                ws.merge_cells(merge_range)
                                
                                # Apply formatting to the merged cell
                                cell = ws.cell(row=current_row, column=span_start)
                                
                                # Apply status-based colors
                                if status == 'Arrived':
                                    cell.fill = colors['arrived']
                                    cell.font = fonts['normal']
                                elif status == 'Confirmed':
                                    cell.fill = colors['confirmed']
                                    cell.font = fonts['normal']
                                elif status == 'Unconfirmed':
                                    cell.fill = colors['unconfirmed']
                                    cell.font = fonts['normal']
                                elif status == 'Maintenance':
                                    cell.fill = colors['maintenance']
                                    cell.font = fonts['normal']
                                elif status == 'Pencil':
                                    cell.fill = colors['pencil']
                                    cell.font = fonts['normal']
                                else:
                                    cell.fill = colors['confirmed']
                                    cell.font = fonts['normal']
                                
                                # Add dart icon for fixed bookings, but keep status color
                                if is_fixed:
                                    cell.value = f"🎯 {self._safe_surname(surname, status)}"
                                else:
                                    cell.value = self._safe_surname(surname, status)
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                                # Enhanced comment with span information
                                span_nights = span_end - span_start
                                comment_text = (f"Res: {res_no}\n"
                                              f"Guest: {surname}\n"
                                              f"Status: {status}\n"
                                              f"Span: {span_nights} nights\n"
                                              f"Total: {res_info['nights']} nights")
                                if is_fixed:
                                    comment_text += f"\n🔒 FIXED - Cannot move"
                                
                                comment = Comment(comment_text, "DefragScript")
                                comment.width = 150
                                comment.height = 100
                                cell.comment = comment
                                
                                col_idx = span_end
                            else:
                                # Single night booking (no merge needed)
                                cell = ws.cell(row=current_row, column=col_idx)
                                
                                # Apply status-based colors
                                if status == 'Arrived':
                                    cell.fill = colors['arrived']
                                    cell.font = fonts['normal']
                                elif status == 'Confirmed':
                                    cell.fill = colors['confirmed']
                                    cell.font = fonts['normal']
                                elif status == 'Unconfirmed':
                                    cell.fill = colors['unconfirmed']
                                    cell.font = fonts['normal']
                                elif status == 'Maintenance':
                                    cell.fill = colors['maintenance']
                                    cell.font = fonts['normal']
                                elif status == 'Pencil':
                                    cell.fill = colors['pencil']
                                    cell.font = fonts['normal']
                                else:
                                    cell.fill = colors['confirmed']
                                    cell.font = fonts['normal']
                                
                                # Add dart icon for fixed bookings, but keep status color
                                if is_fixed:
                                    cell.value = f"🎯 {self._safe_surname(surname, status)}"
                                else:
                                    cell.value = self._safe_surname(surname, status)
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                cell.border = borders['reservation']
                                
                                comment_text = (f"Res: {res_no}\n"
                                              f"Guest: {surname}\n"
                                              f"Status: {status}\n"
                                              f"Nights: {res_info['nights']}")
                                if is_fixed:
                                    comment_text += f"\n🔒 FIXED - Cannot move"
                                
                                comment = Comment(comment_text, "DefragScript")
                                comment.width = 150
                                comment.height = 80
                                cell.comment = comment
                                
                                col_idx += 1
                    else:
                        # Empty cell
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.value = ""
                        col_idx += 1
                
                current_row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        for col_idx in range(2, len(date_range) + 2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 8
        
        # Add legend
        self._add_chart_legend(ws, current_row + 2, constraint_start_date, constraint_end_date)
        
        # Return category importance levels for consolidated analysis
        category_importance_levels = {}
        for date in date_range:
            category_importance_levels[date] = {}
            for category in category_daily_opportunities[date]:
                strategic_score = category_daily_opportunities[date][category].get('strategic_score', 0.0)
                importance_level = self._get_importance_level(strategic_score)
                category_importance_levels[date][category] = importance_level
        
        return category_importance_levels
    
    def _create_suggestions_table_sheet(self, ws, suggestions: List[Dict], property_id: int, 
                                      property_name: str, constraint_start_date, constraint_end_date):
        """Create detailed suggestions table sheet"""
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        normal_font = Font(color="000000")
        
        headers = [
            "Move Order", "Reservation No", "Guest Surname", "Current Unit", 
            "Suggested Unit", "Category", "Status", "Arrive Date", 
            "Depart Date", "Nights", "Improvement Score", "Reason"
        ]
        
        # Create headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Add data rows
        for row_idx, suggestion in enumerate(suggestions, start=2):
            data = [
                suggestion['Sequential_Order'],
                suggestion['Reservation_No'],
                suggestion['Surname'],
                suggestion['Current_Unit'],
                suggestion['Suggested_Unit'],
                suggestion['Category'],
                suggestion['Status'],
                suggestion['Arrive_Date'],
                suggestion['Depart_Date'],
                suggestion['Nights'],
                suggestion['Improvement_Score'],
                suggestion['Reason']
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = normal_font
                
                if col_idx in [1, 2, 7, 8, 9, 10, 11]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # Set column widths
        column_widths = [12, 15, 15, 18, 18, 35, 12, 12, 12, 8, 15, 60]
        for col_idx, width in enumerate(column_widths, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
        
        # Add summary section
        summary_start_row = len(suggestions) + 4
        summary_info = [
            "ANALYSIS SUMMARY:",
            f"Property: {property_name} (ID: {property_id})",
            f"Total Moves Suggested: {len(suggestions)}",
            f"Analysis Period: {constraint_start_date} to {constraint_end_date}",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "IMPLEMENTATION NOTES:",
            "• Apply moves in category-based order (Move Order column - e.g., 1.1, 1.2, 2.1, 2.2)",
            "• Each move considers the effect of previous moves",
            "• Only reservations entirely within analysis period are included",
            "• Fixed reservations are excluded from suggestions",
            "• Higher improvement scores = better defragmentation results",
            "• Real-time data from RMS API at time of analysis"
        ]
        
        for i, info in enumerate(summary_info):
            cell = ws.cell(row=summary_start_row + i, column=1, value=info)
            if info.endswith(":"):
                cell.font = Font(bold=True)
            else:
                cell.font = normal_font
            ws.merge_cells(f'A{summary_start_row + i}:L{summary_start_row + i}')

    def _get_heatmap_color(self, score: float, max_score: float) -> PatternFill:
        """Get heatmap color based on score intensity"""
        if max_score == 0:
            hex_color = "FFE6E6"  # Very light red
        else:
            intensity = min(score / max_score, 1.0)
            if intensity < 0.01:
                hex_color = "FFE6E6"  # Very light red
            elif intensity < 0.3:
                hex_color = "FFB3B3"  # Light red
            elif intensity < 0.6:
                hex_color = "FF6666"  # Medium red
            else:
                hex_color = "CC0000"  # Dark red
        
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    
    def _add_chart_legend(self, ws, start_row: int, constraint_start_date, constraint_end_date):
        """Add legend and instructions to chart"""
        ws.cell(row=start_row, column=1, value="LEGEND & INSTRUCTIONS:").font = Font(bold=True)
        
        legend_items = [
            ("Arrived (Cannot Move)", "5B9BD5"),
            ("Confirmed (Can Move)", "70AD47"), 
            ("Unconfirmed (Can Move)", "FFC000"),
            ("🎯 Fixed (Cannot Move) - Dart icon prefix", "70AD47"),
            ("Maintenance", "800080"),
            ("Pencil", "FF69B4"),
            ("SUGGESTED MOVE", "FF6B6B"),
            ("⬆️ Move Up", "FF6B6B"),
            ("⬇️ Move Down", "FF6B6B")
        ]
        
        for i, (text, color) in enumerate(legend_items):
            row = start_row + i + 1
            legend_cell = ws.cell(row=row, column=1, value=text)
            legend_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            if color in ["FF6B6B", "8B4513", "191970"]:
                legend_cell.font = Font(color="FFFFFF")
        
        instructions_row = start_row + len(legend_items) + 3
        instructions = [
            "INSTRUCTIONS:",
            "• Daily heatmap shows optimization potential for each day",
            "• Horizontal blocks = Multi-night bookings (merged cells)",
            "• Red cells with arrows + numbers = Move suggestions (apply in order)",
            "• ⬆️/⬇️ arrows show move direction in the chart (positioned before number)",
            "• 🎯 Dart icon prefix = Fixed reservations (cannot be moved)",
            "• Hover over cells for detailed information",
            f"• Analysis period: {constraint_start_date} to {constraint_end_date}",
            "• See 'Move Suggestions' sheet for detailed implementation table"
        ]
        
        for i, instruction in enumerate(instructions):
            cell = ws.cell(row=instructions_row + i, column=1, value=instruction)
            if i == 0:
                cell.font = Font(bold=True)
            ws.merge_cells(f'A{instructions_row + i}:F{instructions_row + i}')

    # UTILITY METHODS
    # ===============
    
    def _parse_date(self, date_str: str):
        """Parse date string in format 'DD/MM/YYYY HH:MM'"""
        try:
            return datetime.strptime(date_str.split()[0], '%d/%m/%Y').date()
        except:
            try:
                return datetime.strptime(date_str, '%d/%m/%Y').date()
            except:
                print(f"Warning: Could not parse date: {date_str}")
                return None
    
    def _is_reservation_fixed(self, reservation_row) -> bool:
        """Check if reservation is marked as fixed"""
        if 'Fixed' in reservation_row:
            fixed_value = reservation_row['Fixed']
            if isinstance(fixed_value, bool):
                return fixed_value
            elif isinstance(fixed_value, str):
                return fixed_value.lower() in ['true', '1', 'yes']
        return False
    
    def _safe_surname(self, surname_value, status=None) -> str:
        """Safely handle surname values"""
        if pd.isna(surname_value) or surname_value is None:
            if status == 'Maintenance':
                return "Maint"
            elif status == 'Pencil':
                return "Pencil"
            else:
                return "Unknown"
        
        surname_str = str(surname_value)
        return surname_str[:8] if len(surname_str) > 8 else surname_str
    
    def _get_move_direction_icon(self, current_unit: str, target_unit: str, unit_positions: dict) -> str:
        """Get directional arrow icon based on move direction"""
        if current_unit not in unit_positions or target_unit not in unit_positions:
            return "↔️"  # Default to bidirectional arrow if positions unknown
        
        current_pos = unit_positions[current_unit]
        target_pos = unit_positions[target_unit]
        
        if target_pos < current_pos:
            return "⬆️"  # Moving up in the chart
        elif target_pos > current_pos:
            return "⬇️"  # Moving down in the chart
        else:
            return "↔️"  # Same position (shouldn't happen but safety)
    
    def _get_importance_level(self, strategic_score: float) -> str:
        """Convert strategic score to Low/Medium/High importance level"""
        if strategic_score <= 0:
            return "None"
        elif strategic_score < 0.33:
            return "Low"
        elif strategic_score < 0.67:
            return "Medium"
        else:
            return "High"