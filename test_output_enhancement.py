#!/usr/bin/env python3
"""
Output Enhancement Test Script for Step 5
Tests the holiday-enhanced Excel and email output functionality
"""

import os
import sys
import time
import pandas as pd
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional, Tuple

def test_excel_generator_enhancement():
    """Test Excel generator holiday enhancement methods"""
    print("🧪 Testing Excel Generator Holiday Enhancement")
    print("=" * 60)
    
    try:
        from excel_generator import ExcelGenerator
        
        # Test generator creation
        generator = ExcelGenerator()
        
        # Test holiday methods exist
        assert hasattr(generator, 'create_holiday_enhanced_excel'), "Missing create_holiday_enhanced_excel method"
        assert hasattr(generator, '_add_holiday_moves_sheet'), "Missing _add_holiday_moves_sheet method"
        assert hasattr(generator, '_create_holiday_summary_sheet'), "Missing _create_holiday_summary_sheet method"
        
        print("✅ Excel generator holiday methods: PASSED")
        return generator
        
    except ImportError as e:
        print(f"❌ Import error: {e}")
        return None
    except Exception as e:
        print(f"❌ Creation error: {e}")
        return None

def test_email_sender_enhancement():
    """Test email sender holiday enhancement methods"""
    print("\n🧪 Testing Email Sender Holiday Enhancement")
    print("=" * 60)
    
    try:
        from email_sender import EmailSender
        
        # Test sender creation
        sender = EmailSender()
        
        # Test holiday methods exist
        assert hasattr(sender, 'send_holiday_enhanced_email'), "Missing send_holiday_enhanced_email method"
        assert hasattr(sender, '_create_holiday_enhanced_email_message'), "Missing _create_holiday_enhanced_email_message method"
        assert hasattr(sender, '_create_holiday_enhanced_html_content'), "Missing _create_holiday_enhanced_html_content method"
        
        print("✅ Email sender holiday methods: PASSED")
        return sender
        
    except ImportError as e:
        print(f"❌ Import error: {e}")
        return None
    except Exception as e:
        print(f"❌ Creation error: {e}")
        return None

def test_holiday_moves_sheet_creation(generator):
    """Test holiday moves sheet creation functionality"""
    print("\n🧪 Testing Holiday Moves Sheet Creation")
    print("=" * 60)
    
    if not generator:
        print("❌ No generator available for testing")
        return False
    
    try:
        from openpyxl import Workbook
        
        # Create test workbook
        wb = Workbook()
        
        # Create test holiday moves
        test_holiday_moves = [
            {
                'move_id': 'H1',
                'from_unit': 'Cabin 1',
                'to_unit': 'Cabin 2',
                'from_date': '2025-01-20',
                'to_date': '2025-01-25',
                'guest_name': 'John Doe',
                'improvement_score': 0.85,
                'holiday_period': 'Australia Day 2025',
                'holiday_type': 'Public Holiday',
                'holiday_importance': 'High',
                'reasoning': 'Improves contiguous availability for Australia Day period'
            },
            {
                'move_id': 'H2',
                'from_unit': 'Cabin 3',
                'to_unit': 'Cabin 4',
                'from_date': '2025-01-22',
                'to_date': '2025-01-27',
                'guest_name': 'Jane Smith',
                'improvement_score': 0.75,
                'holiday_period': 'Australia Day 2025',
                'holiday_type': 'Public Holiday',
                'holiday_importance': 'High',
                'reasoning': 'Consolidates bookings for better holiday period optimization'
            }
        ]
        
        # Test holiday moves sheet creation
        generator._add_holiday_moves_sheet(wb, test_holiday_moves, 'Test Property')
        
        # Verify sheet was created
        assert 'Holiday Move Suggestions' in wb.sheetnames, "Holiday moves sheet not created"
        
        # Get the sheet
        ws = wb['Holiday Move Suggestions']
        
        # Verify headers
        expected_headers = [
            'Move ID', 'Property', 'From Unit', 'To Unit', 'From Date', 'To Date',
            'Guest Name', 'Improvement Score', 'Holiday Period', 'Holiday Type',
            'Holiday Importance', 'Reasoning'
        ]
        
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=1, column=col).value
            assert actual_header == expected_header, f"Header mismatch at column {col}: expected {expected_header}, got {actual_header}"
        
        # Verify data rows
        assert ws.cell(row=2, column=1).value == 'H1', "First move ID not found"
        assert ws.cell(row=3, column=1).value == 'H2', "Second move ID not found"
        
        print(f"✅ Holiday moves sheet created with {len(test_holiday_moves)} moves")
        print("✅ Headers and data validation: PASSED")
        return True
        
    except Exception as e:
        print(f"❌ Holiday moves sheet creation error: {e}")
        return False

def test_holiday_summary_sheet_creation(generator):
    """Test holiday summary sheet creation functionality"""
    print("\n🧪 Testing Holiday Summary Sheet Creation")
    print("=" * 60)
    
    if not generator:
        print("❌ No generator available for testing")
        return False
    
    try:
        from openpyxl import Workbook
        
        # Create test workbook
        wb = Workbook()
        
        # Create test holiday data
        test_holiday_data = {
            'holiday_periods': [
                {
                    'name': 'Australia Day 2025',
                    'type': 'Public Holiday',
                    'importance': 'High',
                    'start_date': date(2025, 1, 27),
                    'end_date': date(2025, 1, 27),
                    'extended_start': date(2025, 1, 20),
                    'extended_end': date(2025, 2, 3),
                    'state_code': 'NSW'
                },
                {
                    'name': 'Christmas Day 2025',
                    'type': 'Public Holiday',
                    'importance': 'High',
                    'start_date': date(2025, 12, 25),
                    'end_date': date(2025, 12, 25),
                    'extended_start': date(2025, 12, 18),
                    'extended_end': date(2026, 1, 1),
                    'state_code': 'NSW'
                }
            ]
        }
        
        # Test holiday summary sheet creation
        generator._create_holiday_summary_sheet(wb, test_holiday_data, 'Test Property')
        
        # Verify sheet was created
        assert 'Holiday Summary' in wb.sheetnames, "Holiday summary sheet not created"
        
        # Get the sheet
        ws = wb['Holiday Summary']
        
        # Verify title
        assert ws.cell(row=1, column=1).value == 'HOLIDAY ANALYSIS SUMMARY', "Title not found"
        
        # Verify property info
        assert 'Test Property' in ws.cell(row=3, column=1).value, "Property info not found"
        
        # Verify holiday periods table headers
        expected_headers = ['Holiday Name', 'Type', 'Importance', 'Date', 'Extended Period', 'State']
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=7, column=col).value
            assert actual_header == expected_header, f"Header mismatch at column {col}: expected {expected_header}, got {actual_header}"
        
        # Verify holiday periods data
        assert ws.cell(row=8, column=1).value == 'Australia Day 2025', "First holiday name not found"
        assert ws.cell(row=9, column=1).value == 'Christmas Day 2025', "Second holiday name not found"
        
        print(f"✅ Holiday summary sheet created with {len(test_holiday_data['holiday_periods'])} holiday periods")
        print("✅ Title, headers, and data validation: PASSED")
        return True
        
    except Exception as e:
        print(f"❌ Holiday summary sheet creation error: {e}")
        return False

def test_holiday_enhanced_excel_creation(generator):
    """Test complete holiday-enhanced Excel creation"""
    print("\n🧪 Testing Complete Holiday-Enhanced Excel Creation")
    print("=" * 60)
    
    if not generator:
        print("❌ No generator available for testing")
        return False
    
    try:
        # Create test data
        test_reservations = [
            {
                'Arrive': '20/01/2025',
                'Depart': '25/01/2025',
                'Guest Name': 'John Doe',
                'Unit/Site': 'Cabin 1',
                'Category': 'Cabin',
                'Res No': '12345',
                'Surname': 'Doe',
                'Status': 'Confirmed',
                'Fixed': 'False',
                'Nights': '5'
            }
        ]
        
        test_inventory = [
            {
                'Category': 'Cabin',
                'Unit/Site': 'Cabin 1',
                'Type': 'Cabin'
            },
            {
                'Category': 'Cabin',
                'Unit/Site': 'Cabin 2',
                'Type': 'Cabin'
            }
        ]
        
        test_regular_suggestions = [
            {
                'Sequential_Order': '1.1',
                'Reservation_No': '12345',
                'Surname': 'Doe',
                'Current_Unit': 'Cabin 1',
                'Suggested_Unit': 'Cabin 2',
                'Category': 'Cabin',
                'Status': 'Confirmed',
                'Arrive_Date': '20/01/2025',
                'Depart_Date': '25/01/2025',
                'Nights': 5,
                'Improvement_Score': 0.85,
                'Reason': 'Improves contiguous availability'
            }
        ]
        
        test_holiday_suggestions = [
            {
                'move_id': 'H1',
                'from_unit': 'Cabin 1',
                'to_unit': 'Cabin 2',
                'from_date': '20/01/2025',
                'to_date': '25/01/2025',
                'guest_name': 'John Doe',
                'improvement_score': 0.85,
                'holiday_period': 'Australia Day 2025',
                'holiday_type': 'Public Holiday',
                'holiday_importance': 'High',
                'reasoning': 'Improves contiguous availability for Australia Day period',
                'Nights': 5
            }
        ]
        
        test_holiday_data = {
            'holiday_periods': [
                {
                    'name': 'Australia Day 2025',
                    'type': 'Public Holiday',
                    'importance': 'High',
                    'start_date': date(2025, 1, 27),
                    'end_date': date(2025, 1, 27),
                    'extended_start': date(2025, 1, 20),
                    'extended_end': date(2025, 2, 3),
                    'state_code': 'NSW'
                }
            ]
        }
        
        reservations_df = pd.DataFrame(test_reservations)
        inventory_df = pd.DataFrame(test_inventory)
        
        # Test file path
        test_file = 'test_holiday_enhanced_output.xlsx'
        
        # Test holiday-enhanced Excel creation
        success, category_importance = generator.create_holiday_enhanced_excel(
            reservations_df, inventory_df, test_regular_suggestions, test_holiday_suggestions,
            123, 'Test Property', test_holiday_data, date(2025, 1, 20), date(2025, 2, 3), test_file
        )
        
        assert success == True, "Holiday-enhanced Excel creation failed"
        assert os.path.exists(test_file), "Holiday-enhanced Excel file not created"
        
        # Verify file can be opened and has expected sheets
        from openpyxl import load_workbook
        wb = load_workbook(test_file)
        
        expected_sheets = ['Visual Chart', 'Move Suggestions', 'Holiday Move Suggestions', 'Holiday Summary']
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames, f"Expected sheet '{sheet_name}' not found"
        
        print(f"✅ Holiday-enhanced Excel file created: {test_file}")
        print(f"✅ All expected sheets present: {expected_sheets}")
        
        # Clean up test file
        os.remove(test_file)
        print("✅ Test file cleaned up")
        
        return True
        
    except Exception as e:
        print(f"❌ Holiday-enhanced Excel creation error: {e}")
        return False

def test_holiday_enhanced_email_content(sender):
    """Test holiday-enhanced email content creation"""
    print("\n🧪 Testing Holiday-Enhanced Email Content")
    print("=" * 60)
    
    if not sender:
        print("❌ No sender available for testing")
        return False
    
    try:
        # Create test data
        test_regular_suggestions = [
            {
                'Sequential_Order': '1.1',
                'Reservation_No': '12345',
                'Surname': 'Doe',
                'Current_Unit': 'Cabin 1',
                'Suggested_Unit': 'Cabin 2',
                'Improvement_Score': 0.85
            }
        ]
        
        test_holiday_suggestions = [
            {
                'move_id': 'H1',
                'guest_name': 'John Doe',
                'from_unit': 'Cabin 1',
                'to_unit': 'Cabin 2',
                'holiday_period': 'Australia Day 2025',
                'holiday_importance': 'High',
                'improvement_score': 0.85
            }
        ]
        
        test_holiday_data = {
            'holiday_periods': [
                {
                    'name': 'Australia Day 2025',
                    'type': 'Public Holiday',
                    'importance': 'High'
                }
            ]
        }
        
        # Test HTML content creation
        html_content = sender._create_holiday_enhanced_html_content(
            'Test Property', 123, test_regular_suggestions, test_holiday_suggestions,
            test_holiday_data, date(2025, 1, 20), date(2025, 2, 3), True
        )
        
        # Verify HTML content contains expected elements
        assert 'Holiday-Enhanced RMS Defragmentation Analysis' in html_content, "Email title not found"
        assert 'Test Property' in html_content, "Property name not found"
        assert 'Holiday Move Recommendations' in html_content, "Holiday moves section not found"
        assert 'Regular Move Recommendations' in html_content, "Regular moves section not found"
        assert 'H1' in html_content, "Holiday move ID not found"
        assert 'Australia Day 2025' in html_content, "Holiday period not found"
        assert 'High' in html_content, "Holiday importance not found"
        
        print("✅ Holiday-enhanced email HTML content created successfully")
        print("✅ All expected content elements present")
        return True
        
    except Exception as e:
        print(f"❌ Holiday-enhanced email content error: {e}")
        return False

def test_email_message_creation(sender):
    """Test holiday-enhanced email message creation"""
    print("\n🧪 Testing Holiday-Enhanced Email Message Creation")
    print("=" * 60)
    
    if not sender:
        print("❌ No sender available for testing")
        return False
    
    try:
        # Create test data
        test_regular_suggestions = []
        test_holiday_suggestions = []
        test_holiday_data = {'holiday_periods': []}
        
        # Test email message creation
        msg = sender._create_holiday_enhanced_email_message(
            'Test Property', 123, test_regular_suggestions, test_holiday_suggestions,
            test_holiday_data, date(2025, 1, 20), date(2025, 2, 3), True, 'test@example.com'
        )
        
        # Verify email message structure
        assert msg['From'] == sender.sender_email, "From address not set correctly"
        assert msg['To'] == 'test@example.com', "To address not set correctly"
        assert 'Holiday-Enhanced Defragmentation Analysis' in msg['Subject'], "Subject not set correctly"
        
        # Verify HTML content is attached
        html_parts = [part for part in msg.walk() if part.get_content_type() == 'text/html']
        assert len(html_parts) > 0, "HTML content not attached"
        
        print("✅ Holiday-enhanced email message created successfully")
        print("✅ Email headers and content structure validated")
        return True
        
    except Exception as e:
        print(f"❌ Holiday-enhanced email message creation error: {e}")
        return False

def run_output_enhancement_tests():
    """Run all output enhancement tests"""
    print("🚀 Running Output Enhancement Tests for Step 5")
    print("=" * 70)
    print(f"Test Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    try:
        # Test 1: Excel generator enhancement
        generator = test_excel_generator_enhancement()
        if not generator:
            return False
        
        # Test 2: Email sender enhancement
        sender = test_email_sender_enhancement()
        if not sender:
            return False
        
        # Test 3: Holiday moves sheet creation
        if not test_holiday_moves_sheet_creation(generator):
            return False
        
        # Test 4: Holiday summary sheet creation
        if not test_holiday_summary_sheet_creation(generator):
            return False
        
        # Test 5: Complete holiday-enhanced Excel creation
        if not test_holiday_enhanced_excel_creation(generator):
            return False
        
        # Test 6: Holiday-enhanced email content
        if not test_holiday_enhanced_email_content(sender):
            return False
        
        # Test 7: Email message creation
        if not test_email_message_creation(sender):
            return False
        
        print("\n" + "=" * 70)
        print("🎉 ALL OUTPUT ENHANCEMENT TESTS PASSED!")
        print("✅ Holiday-enhanced Excel and email functionality is working correctly")
        print("📋 Next: Step 6 - Implement Main Application Integration")
        
        return True
        
    except Exception as e:
        print(f"\n❌ UNEXPECTED ERROR: {e}")
        return False

if __name__ == "__main__":
    success = run_output_enhancement_tests()
    sys.exit(0 if success else 1)
